import openpyxl
import pandas as pd
from pathlib import Path

def extract_excel_data(file_path):
    """
    Extract all data from an Excel file (.xlsx)
    
    Args:
        file_path (str): Path to the .xlsx file
    
    Returns:
        dict: Dictionary containing data from all sheets
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        
        # Dictionary to store all sheets' data
        all_data = {}
        
        print(f"Processing file: {file_path}")
        print(f"Number of sheets: {len(workbook.sheetnames)}\n")
        
        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            print(f"--- Sheet: {sheet_name} ---")
            print(f"Dimensions: {sheet.dimensions}")
            
            # Extract data from the sheet
            sheet_data = []
            for row in sheet.iter_rows(values_only=True):
                sheet_data.append(row)
            
            all_data[sheet_name] = sheet_data
            
            # Display first few rows
            print(f"First 5 rows:")
            for i, row in enumerate(sheet_data[:5]):
                print(f"  Row {i+1}: {row}")
            print()
        
        return all_data
    
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return None
    except Exception as e:
        print(f"Error: {e}")
        return None


def extract_excel_with_pandas(file_path):
    """
    Alternative method using pandas to extract Excel data
    
    Args:
        file_path (str): Path to the .xlsx file
    
    Returns:
        dict: Dictionary of DataFrames, one per sheet
    """
    try:
        # Read all sheets into a dictionary of DataFrames
        all_sheets = pd.read_excel(file_path, sheet_name=None)
        
        print(f"Processing file: {file_path}")
        print(f"Number of sheets: {len(all_sheets)}\n")
        
        # Display info for each sheet
        for sheet_name, df in all_sheets.items():
            print(f"--- Sheet: {sheet_name} ---")
            print(f"Shape: {df.shape} (rows, columns)")
            print(f"Columns: {list(df.columns)}")
            print("\nFirst 5 rows:")
            print(df.head())
            print("\n")
        
        return all_sheets
    
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return None
    except Exception as e:
        print(f"Error: {e}")
        return None


if __name__ == "__main__":
    # Example usage
    file_path = "ReportTester-263254895.xlsx"  # Change this to your file path
    
    print("=" * 60)
    print("Method 1: Using openpyxl")
    print("=" * 60)
    data = extract_excel_data(file_path)
    
    print("\n" + "=" * 60)
    print("Method 2: Using pandas (recommended)")
    print("=" * 60)
    df_data = extract_excel_with_pandas(file_path)
    
    # Optional: Save extracted data
    if df_data:
        # Export each sheet to CSV
        for sheet_name, df in df_data.items():
            output_file = f"{sheet_name}.csv"
            df.to_csv(output_file, index=False)
            print(f"Exported '{sheet_name}' to {output_file}")